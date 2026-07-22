"""Workbook vivo per FinanceBoost, costruito solo da input riconciliati."""
from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any


def render_finance_workbook(inputs: dict, path: Path, pack: dict | None = None) -> Path:
    """Workbook FinanceBoost. Coi bilanci → fogli di analisi di bilancio; col
    pacchetto consulenziale (pack) → in più i fogli di tesoreria VIVI (forecast,
    aging, KPI, simulazioni, piano, registro decisioni). Senza bilanci ma con
    pack → workbook di sola tesoreria (il caso liquidità non ha bilanci)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = [r for r in (inputs.get("bilanci") or []) if isinstance(r, dict)]
    if not rows and pack is None:
        raise ValueError("bilanci mancanti")
    if not rows:
        wb = Workbook()
        wb.remove(wb.active)
        _append_treasury_sheets(wb, inputs, pack)
        append_decision_sheets(wb, pack)
        _finalize(wb, path)
        return path
    latest = max(rows, key=lambda r: int(r.get("anno") or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Input bilancio"
    ws.append(["FinanceBoost - input riconciliati", None, "Fonte"])
    ws["A1"].font = Font(bold=True, size=14, color="063B36")
    ws.append(["Voce", "Valore", "Nota"])
    labels = [
        ("Ricavi", "ricavi"), ("EBITDA", "ebitda"), ("EBIT", "reddito_operativo"),
        ("Utile netto", "utile_netto"), ("Totale attivo", "totale_attivo"),
        ("Attivo corrente", "attivo_corrente"), ("Passivo corrente", "passivo_corrente"),
        ("Rimanenze", "rimanenze"), ("Patrimonio netto", "patrimonio_netto"),
        ("Debiti finanziari", "debiti_finanziari"),
    ]
    input_rows: dict[str, int] = {}
    for label, key in labels:
        ws.append([label, latest.get(key), f"Esercizio {latest.get('anno', '')}"])
        input_rows[key] = ws.max_row
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="0C7A6F"); cell.font = Font(color="FFFFFF", bold=True)
    for row in range(3, ws.max_row + 1):
        ws.cell(row, 2).number_format = '#,##0.00;[Red](#,##0.00);-'
        ws.cell(row, 2).font = Font(name="Arial", color="0000FF")
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 18; ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A3"

    ind = wb.create_sheet("Indici")
    ind.append(["Indice", "Valore", "Descrizione", "Stato dati"])
    for c in ind[1]:
        c.fill = PatternFill("solid", fgColor="0C7A6F"); c.font = Font(color="FFFFFF", bold=True)
    ref = lambda key: f"'Input bilancio'!B{input_rows[key]}"
    formulas = [
        ("Debt / Equity", f'=IFERROR({ref("debiti_finanziari")}/{ref("patrimonio_netto")},"")', "Debiti finanziari / PN"),
        ("ROE", f'=IFERROR({ref("utile_netto")}/{ref("patrimonio_netto")},"")', "Utile netto / PN"),
        ("ROS", f'=IFERROR({ref("reddito_operativo")}/{ref("ricavi")},"")', "EBIT / Ricavi"),
        ("ROI proxy", f'=IFERROR({ref("reddito_operativo")}/{ref("totale_attivo")},"")', "EBIT / Totale attivo"),
        ("EBITDA margin", f'=IFERROR({ref("ebitda")}/{ref("ricavi")},"")', "EBITDA / Ricavi"),
        ("Current ratio", f'=IFERROR({ref("attivo_corrente")}/{ref("passivo_corrente")},"")', "Attivo corrente / Passivo corrente"),
        ("Quick ratio", f'=IFERROR(({ref("attivo_corrente")}-{ref("rimanenze")})/{ref("passivo_corrente")},"")', "(AC - rimanenze) / PC"),
        ("CCN", f'=IF(OR({ref("attivo_corrente")}="",{ref("passivo_corrente")}=""),"",{ref("attivo_corrente")}-{ref("passivo_corrente")})', "Attivo corrente - Passivo corrente"),
    ]
    for name, formula, description in formulas:
        ind.append([name, formula, description, "formula viva; vuoto se input mancanti"])
        ind.cell(ind.max_row, 2).number_format = "0.00%" if name in {"ROE", "ROS", "ROI proxy", "EBITDA margin"} else "0.00"
    ind.column_dimensions["A"].width = 24; ind.column_dimensions["B"].width = 16
    ind.column_dimensions["C"].width = 36; ind.column_dimensions["D"].width = 16
    ind.freeze_panes = "A2"

    sc = wb.create_sheet("Scenari")
    sc.append(["Scenario", "Crescita ricavi (input)", "Variazione margine EBITDA (input)", "Ricavi proiettati", "EBITDA proiettato"])
    for c in sc[1]:
        c.fill = PatternFill("solid", fgColor="0C7A6F"); c.font = Font(color="FFFFFF", bold=True)
    for scenario in ("Base", "Ottimistico", "Pessimistico"):
        sc.append([scenario, None, None, f'=IF(B{sc.max_row+1}="","",{ref("ricavi")}*(1+B{sc.max_row+1}))',
                   f'=IF(OR(B{sc.max_row+1}="",C{sc.max_row+1}=""),"",D{sc.max_row+1}*({ref("ebitda")}/{ref("ricavi")}+C{sc.max_row+1}))'])
    for row in range(2, 5):
        sc.cell(row, 2).font = Font(color="0000FF"); sc.cell(row, 3).font = Font(color="0000FF")
        sc.cell(row, 2).fill = PatternFill("solid", fgColor="FFF2CC"); sc.cell(row, 3).fill = PatternFill("solid", fgColor="FFF2CC")
        sc.cell(row, 2).number_format = "0.0%"; sc.cell(row, 3).number_format = "0.0%"
        sc.cell(row, 4).number_format = '#,##0.00'; sc.cell(row, 5).number_format = '#,##0.00'
    for col, width in zip("ABCDE", (18, 24, 34, 22, 22)):
        sc.column_dimensions[col].width = width
    sc.freeze_panes = "A2"
    sc["A6"] = "Le celle gialle/blu sono assunzioni utente: nessuno scenario viene inventato dal sistema."
    sc.merge_cells("A6:E6"); sc["A6"].alignment = Alignment(wrap_text=True)

    if pack:
        _append_treasury_sheets(wb, inputs, pack)
        append_decision_sheets(wb, pack)

    wb.calculation.forceFullCalc = True
    _finalize(wb, path)
    return path


def append_decision_sheets(wb, pack: dict | None) -> None:
    """Fogli decisionali dal pacchetto consulenziale (#14 review deliverable — l'Excel è
    uno strumento operativo, non un dump): «Confronto opzioni» (matrice decisionale con
    colonna Decisione da compilare) e «Parametri da definire» (celle EDITABILI gialle per
    i dati mancanti/da personalizzare — il PDF li cita come 'Parametro da definire').
    No-op senza pack o senza dati (mai inventa righe)."""
    if not isinstance(pack, dict):
        return
    from openpyxl.styles import Alignment, Font, PatternFill

    _EDIT = PatternFill("solid", fgColor="FFF2CC")
    _HDRF = PatternFill("solid", fgColor="0C7A6F")

    conf = pack.get("confronto_soluzioni") or {}
    opzioni = [o for o in (conf.get("opzioni") or []) if isinstance(o, dict)]
    if opzioni:
        ws = wb.create_sheet("Confronto opzioni")
        ws.append(["Opzione", "Costi", "Tempi", "Rischi", "Complessità",
                   "Quando sceglierla", "Quando evitarla", "Decisione (compila)"])
        for c in ws[1]:
            c.fill = _HDRF; c.font = Font(color="FFFFFF", bold=True)
        for o in opzioni:
            ws.append([str(o.get(k) or "") for k in
                       ("opzione", "costi", "tempi", "rischi", "complessita",
                        "quando_sceglierla", "quando_evitarla")] + [None])
            dcell = ws.cell(ws.max_row, 8)
            dcell.fill = _EDIT; dcell.font = Font(color="0000FF")
        for col, width in zip("ABCDEFGH", (30, 26, 16, 26, 12, 34, 34, 18)):
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        if conf.get("conclusione_motivata"):
            ws.append([]); ws.append([f"Conclusione motivata: {conf['conclusione_motivata']}"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=8)
            ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True)

    dati = [str(d) for d in (pack.get("dati_da_raccogliere") or []) if str(d).strip()]
    kpis = [k for k in (pack.get("kpi_da_misurare") or []) if isinstance(k, dict)]
    if dati or kpis:
        ws = wb.create_sheet("Parametri da definire")
        ws.append(["Parametro / KPI", "Perché serve", "Valore (compila)", "Responsabile (compila)"])
        for c in ws[1]:
            c.fill = _HDRF; c.font = Font(color="FFFFFF", bold=True)
        for d in dati[:12]:
            ws.append([d, "dato mancante per completare la misurazione", None, None])
        for k in kpis[:10]:
            ws.append([str(k.get("kpi") or ""), str(k.get("perche") or ""), None, None])
        for row in ws.iter_rows(min_row=2):
            row[2].fill = _EDIT; row[2].font = Font(color="0000FF")
            row[3].fill = _EDIT; row[3].font = Font(color="0000FF")
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col, width in zip("ABCD", (36, 44, 18, 20)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"


def _finalize(wb, path: Path) -> None:
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Arial"
                cell.font = font
    wb.calculation.fullCalcOnLoad = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _append_treasury_sheets(wb, inputs: dict, pack: dict) -> None:
    """Fogli tesoreria del pacchetto consulenziale (spec §10): il proseguimento
    operativo del report — formule vive, celle input evidenziate, mai zeri finti."""
    from openpyxl.styles import Alignment, Font, PatternFill

    def hdr(ws, r=1):
        for c in ws[r]:
            if c.value is not None:
                c.fill = PatternFill("solid", fgColor="0C7A6F")
                c.font = Font(color="FFFFFF", bold=True)

    def widths(ws, ww):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(ww, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    from .insight import Facts
    f = Facts(inputs)
    inc = f.get("incassi_mese")
    usc = f.get("uscite_mese")
    sc = f.get("scoperto") or 0.0

    # 1) Forecast 13 settimane — VIVO: cambi le celle gialle, tutto si ricalcola.
    ws = wb.create_sheet("Forecast 13 settimane")
    ws.append(["Parametro", "Valore", "Nota"])
    hdr(ws)
    ws.append(["Incassi medi mese", inc if inc is not None else "Da rilevare",
               "cella INPUT: aggiorna col dato reale"])
    ws.append(["Uscite medie mese", usc if usc is not None else "Da rilevare",
               "cella INPUT"])
    ws.append(["Saldo iniziale (−scoperto)", -sc, "cella INPUT"])
    for r in (2, 3, 4):
        ws.cell(r, 2).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(r, 2).font = Font(color="0000FF")
    ws.append([])
    ws.append(["Settimana", "Entrate", "Uscite", "Saldo progressivo"])
    hdr(ws, 6)
    for w in range(1, 14):
        r = ws.max_row + 1
        prev = "B4" if w == 1 else f"D{r - 1}"
        ws.append([w, '=IF(ISNUMBER($B$2),$B$2/4.33,"")',
                   '=IF(ISNUMBER($B$3),$B$3/4.33,"")',
                   f'=IF(AND(ISNUMBER(B{r}),ISNUMBER(C{r})),{prev}+B{r}-C{r},"")'])
        for col in (2, 3, 4):
            ws.cell(r, col).number_format = '#,##0'
    widths(ws, (24, 16, 16, 18))
    ws.freeze_panes = "A7"
    ws.append([])
    ws.append(["Scenari: duplica il foglio e varia le celle gialle "
               "(prudente: entrate −15%, uscite +5% · critico: entrate −30%)."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="8A7A55")

    # 2) Aging crediti — template da compilare (il dato per-cliente non è stato fornito).
    ws = wb.create_sheet("Aging crediti")
    ws.append(["Cliente", "Importo", "Data fattura", "Scadenza", "Giorni di ritardo",
               "Stato sollecito", "Prossima azione", "Data azione"])
    hdr(ws)
    for _ in range(12):
        ws.append([None] * 8)
    widths(ws, (26, 14, 13, 13, 14, 16, 26, 12))
    ws.freeze_panes = "A2"

    # 3) Piano incassi — settimana per settimana, collegabile all'aging.
    ws = wb.create_sheet("Piano incassi")
    ws.append(["Settimana", "Cliente", "Importo atteso", "Incassato", "Delta", "Note"])
    hdr(ws)
    for _ in range(12):
        r = ws.max_row + 1
        ws.append([None, None, None, None,
                   f'=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),D{r}-C{r},"")', None])
    widths(ws, (12, 26, 15, 13, 12, 26))
    ws.freeze_panes = "A2"

    # 4) KPI tesoreria — dagli insight derivati (stessi numeri del PDF).
    ws = wb.create_sheet("KPI tesoreria")
    ws.append(["KPI", "Valore", "Unità", "Formula", "Dati usati", "Periodicità", "Owner"])
    hdr(ws)
    for i in (pack.get("insight_derivati") or []):
        ws.append([i.get("titolo"), i.get("valore"), i.get("unita") or "",
                   i.get("formula") or "", ", ".join(i.get("dati_usati") or []),
                   "Mensile", DA_RILEVARE])
    if not (pack.get("insight_derivati") or []):
        ws.append(["Nessun KPI derivabile dai dati forniti", DA_RILEVARE, "", "", "", "", ""])
    widths(ws, (30, 14, 10, 40, 30, 12, 14))
    ws.freeze_panes = "A2"

    # 5) Simulazioni — what-if del report, con dati e calcolo.
    ws = wb.create_sheet("Simulazioni")
    ws.append(["Domanda", "Risultato", "Calcolo", "Dati usati"])
    hdr(ws)
    for s in (pack.get("simulazioni") or []):
        ws.append([s.get("domanda"), s.get("risultato"), s.get("calcolo"),
                   ", ".join(s.get("dati_usati") or [])])
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True)
    widths(ws, (44, 34, 30, 26))
    ws.freeze_panes = "A2"

    # 6) Piano azione — dalle raccomandazioni (owner/cadenza dal motore decisionale).
    ws = wb.create_sheet("Piano azione")
    ws.append(["Azione", "Chi", "Quando", "Cadenza", "KPI generati", "Stato"])
    hdr(ws)
    for r_ in (pack.get("raccomandazioni_operative") or []):
        op = r_.get("operativo") or {}
        ws.append([r_.get("titolo"), op.get("chi"), op.get("quando"), op.get("cadenza"),
                   ", ".join(op.get("kpi_generati") or []), "Da avviare"])
    widths(ws, (40, 24, 20, 24, 34, 12))
    ws.freeze_panes = "A2"

    # 7) Registro decisioni — traccia di chi decide cosa (disciplina del metodo).
    ws = wb.create_sheet("Registro decisioni")
    ws.append(["Data", "Decisione", "Chi ha deciso", "Su quali dati", "Esito atteso",
               "Verifica (data)", "Esito reale"])
    hdr(ws)
    for _ in range(10):
        ws.append([None] * 7)
    widths(ws, (12, 40, 18, 30, 26, 14, 26))
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# ControlBoost: cruscotto commesse operativo (spec §11).
# Il foglio KPI è popolato dalla STESSA estrazione usata dal PDF
# (quality_gate.extract_kpis) → coerenza PDF/Excel per costruzione.
# Dato mancante → "Da rilevare", mai 0. Le strutture (stati, RACI, checklist)
# sono TEMPLATE PROPOSTI, marcati come tali: nessun dato cliente inventato.
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FILL = "0C7A6F"
_INPUT_FILL = "FFF2CC"
DA_RILEVARE = "Da rilevare"

# Costanti condivise con il pacchetto consulenziale (fonte unica → PDF/Excel coerenti).
from .consulting import (CHECKLIST_FASI as _CHECKLIST_FASI,
                         RACI_ATTIVITA as _RACI_ATTIVITA,
                         RACI_RUOLI as _RACI_RUOLI,
                         STATI_STANDARD as _STATI_STANDARD)


def _hdr(ws, row_idx: int = 1):
    from openpyxl.styles import Font, PatternFill
    for c in ws[row_idx]:
        if c.value is not None:
            c.fill = PatternFill("solid", fgColor=_HDR_FILL)
            c.font = Font(color="FFFFFF", bold=True)


def _widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def render_control_workbook(deliverable: dict, inputs: dict, path: Path) -> tuple[Path, dict]:
    """Costruisce il cruscotto commesse. Ritorna (path, mapping) dove mapping =
    {reportSectionId: {"sheet": nome_foglio, "range": intervallo}} (spec §11)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from .quality_gate import extract_kpis

    kpis = extract_kpis(deliverable)
    azienda = str(inputs.get("azienda") or inputs.get("ragione_sociale") or "")
    periodo = f"{inputs.get('mese', '')} {inputs.get('anno', '')}".strip()
    mapping: dict[str, dict] = {}

    wb = Workbook()

    # 1) Registro commesse — template vuoto (mai dati inventati).
    ws = wb.active
    ws.title = "Registro commesse"
    ws.append([f"Registro commesse — {azienda}".strip(" —"), None, None, None,
               "Compilare: una riga per commessa. Nessun dato è precompilato."])
    ws["A1"].font = Font(bold=True, size=13, color="063B36")
    ws.append(["ID", "Nome commessa", "Cliente", "Owner (PM)", "Stato", "Priorità",
               "Data apertura", "Prossima azione", "Data prossima azione", "Note"])
    _hdr(ws, 2)
    for _ in range(12):
        ws.append([None] * 10)
    _widths(ws, (10, 30, 24, 18, 16, 10, 14, 30, 16, 30))
    ws.freeze_panes = "A3"
    mapping["registro_commesse"] = {"sheet": ws.title, "range": "A2:J14"}

    # 2) Dizionario stati — proposta standard (marcata come tale).
    ws = wb.create_sheet("Dizionario stati")
    ws.append(["Stato", "Definizione", "Criterio di ingresso", "Criterio di uscita"])
    _hdr(ws)
    for row in _STATI_STANDARD:
        ws.append(list(row))
    ws.append([])
    ws.append(["Proposta iniziale di stati standard: da validare con la direzione prima dell'adozione."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="8A7A55")
    _widths(ws, (18, 44, 34, 34))
    ws.freeze_panes = "A2"
    mapping["stati_commessa"] = {"sheet": ws.title, "range": f"A1:D{len(_STATI_STANDARD) + 1}"}

    # 3) Registro blocchi.
    ws = wb.create_sheet("Registro blocchi")
    ws.append(["ID", "Commessa", "Descrizione blocco", "Motivazione", "Data inizio",
               "Owner sblocco", "Prossima azione", "Data prossima azione", "Stato"])
    _hdr(ws)
    for _ in range(10):
        ws.append([None] * 9)
    _widths(ws, (8, 22, 34, 30, 12, 16, 30, 16, 12))
    ws.freeze_panes = "A2"
    mapping["registro_blocchi"] = {"sheet": ws.title, "range": "A1:I11"}

    # 4) KPI — dalla stessa estrazione del PDF (coerenza per costruzione).
    ws = wb.create_sheet("KPI")
    ws.append(["KPI", "Valore attuale", "Unità", "Target", "Stato", "Scost. %",
               "Formula", "Fonte", "Periodicità", "Owner", "Ultimo aggiornamento"])
    _hdr(ws)
    for k in kpis:
        ws.append([
            k["nome"],
            k["valore"] if k["valore"] is not None else DA_RILEVARE,
            k["unita"] or "",
            k["target"] if k["target"] is not None else DA_RILEVARE,
            (k["semaforo"] or "").capitalize() or DA_RILEVARE,
            None,
            k["formula"] or DA_RILEVARE,
            k["fonte"] or DA_RILEVARE,
            "Mensile",
            DA_RILEVARE,          # owner del KPI: da assegnare, mai inventato
            periodo or DA_RILEVARE,
        ])
        r = ws.max_row
        ws.cell(r, 6).value = (f'=IF(OR(B{r}="{DA_RILEVARE}",D{r}="{DA_RILEVARE}",D{r}=0),"",'
                               f'(B{r}/D{r}-1))')
        ws.cell(r, 6).number_format = "0.0%"
    if not kpis:
        ws.append(["Nessun KPI disponibile nel report", DA_RILEVARE, "", DA_RILEVARE,
                   "", None, "", "", "", "", ""])
    _widths(ws, (28, 14, 8, 12, 10, 10, 34, 30, 12, 14, 18))
    ws.freeze_panes = "A2"
    mapping["kpi"] = {"sheet": ws.title, "range": f"A1:K{ws.max_row}"}

    # 5) Dashboard — formule vive sul foglio KPI (niente valori copiati).
    ws = wb.create_sheet("Dashboard")
    ws.append([f"Dashboard direzionale — {periodo}".strip(" —")])
    ws["A1"].font = Font(bold=True, size=13, color="063B36")
    last = max(len(kpis) + 1, 2)   # ultima riga dati del foglio KPI
    ws.append(["KPI totali", f"=COUNTA(KPI!A2:A{last})"])
    ws.append(["KPI in stato Rosso", f'=COUNTIF(KPI!E2:E{last},"Rosso")'])
    ws.append(["KPI in stato Giallo", f'=COUNTIF(KPI!E2:E{last},"Giallo")'])
    ws.append(["KPI in stato Verde", f'=COUNTIF(KPI!E2:E{last},"Verde")'])
    ws.append(["KPI da rilevare", f'=COUNTIF(KPI!B2:B{last},"{DA_RILEVARE}")'])
    ws.append([])
    ws.append(["Si aggiorna da sola quando compili il foglio KPI. Non inserire valori a mano qui."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="8A7A55")
    _widths(ws, (26, 14))
    mapping["dashboard"] = {"sheet": ws.title, "range": "A2:B7"}

    # 6) Piano 30-60-90 — dalle azioni del report se presenti, senza inventare date.
    ws = wb.create_sheet("Piano 30-60-90")
    ws.append(["Orizzonte", "Azione", "Owner", "Scadenza", "Stato", "Note"])
    _hdr(ws)
    azioni = [a for a in (deliverable.get("azioni") or deliverable.get("piano_azione") or [])
              if isinstance(a, (str, dict))]
    horizons = ("0-30 giorni", "31-60 giorni", "61-90 giorni")
    for i, a in enumerate(azioni[:9]):
        testo = a if isinstance(a, str) else str(a.get("azione") or a.get("titolo") or "")
        ws.append([horizons[min(i // 3, 2)], testo, DA_RILEVARE, DA_RILEVARE, "Da avviare", ""])
    if not azioni:
        for h in horizons:
            ws.append([h, None, None, None, None, None])
    _widths(ws, (14, 52, 16, 12, 12, 26))
    ws.freeze_panes = "A2"
    mapping["piano_30_60_90"] = {"sheet": ws.title, "range": f"A1:F{ws.max_row}"}

    # 7) RACI — attività standard x ruoli PROPOSTI (nessun nome di persona).
    ws = wb.create_sheet("RACI")
    ws.append(["Attività \\ Ruolo"] + [f"{r} (proposto)" for r in _RACI_RUOLI])
    _hdr(ws)
    for att in _RACI_ATTIVITA:
        ws.append([att] + [None] * len(_RACI_RUOLI))
    ws.append([])
    ws.append(["Compilare con R (Responsible), A (Accountable), C (Consulted), I (Informed). "
               "Una sola A per riga. I ruoli sono proposti: adattarli all'organigramma reale."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, color="8A7A55")
    _widths(ws, (26,) + (16,) * len(_RACI_RUOLI))
    ws.freeze_panes = "B2"
    mapping["raci"] = {"sheet": ws.title,
                       "range": f"A1:{chr(ord('A') + len(_RACI_RUOLI))}{len(_RACI_ATTIVITA) + 1}"}

    # 8) Checklist per fase.
    ws = wb.create_sheet("Checklist")
    ws.append(["Fase", "Voce di controllo", "Fatto (Sì/No)", "Note"])
    _hdr(ws)
    for fase, voci in _CHECKLIST_FASI:
        for voce in voci:
            ws.append([fase, voce, None, None])
    _widths(ws, (16, 48, 14, 30))
    ws.freeze_panes = "A2"
    mapping["checklist"] = {"sheet": ws.title, "range": f"A1:D{ws.max_row}"}

    # 9) Log aggiornamenti.
    ws = wb.create_sheet("Log aggiornamenti")
    ws.append(["Data", "Chi", "Cosa è cambiato", "Note"])
    _hdr(ws)
    for _ in range(10):
        ws.append([None] * 4)
    _widths(ws, (12, 16, 48, 30))
    ws.freeze_panes = "A2"
    mapping["log_aggiornamenti"] = {"sheet": ws.title, "range": "A1:D11"}

    # 10) Istruzioni.
    ws = wb.create_sheet("Istruzioni")
    istruzioni = [
        ("Registro commesse", "Una riga per commessa. Stato SOLO dal Dizionario stati. "
                              "Owner unico e data prossima azione sempre compilati."),
        ("Dizionario stati", "Stati standard proposti con criteri di ingresso/uscita. "
                             "Validarli con la direzione, poi non modificarli più."),
        ("Registro blocchi", "Ogni commessa Bloccata DEVE avere una riga qui, con motivazione "
                             "e owner dello sblocco."),
        ("KPI", "Valori del report. 'Da rilevare' = dato non ancora disponibile: sostituirlo "
                "col dato reale appena misurato. Mai inserire 0 al posto di un dato mancante."),
        ("Dashboard", "Si aggiorna automaticamente dal foglio KPI. Non scrivere qui."),
        ("Piano 30-60-90", "Azioni del report distribuite sugli orizzonti. Assegnare owner e scadenze."),
        ("RACI", "Compilare R/A/C/I per ogni attività. Una sola A per riga."),
        ("Checklist", "Spuntare le voci a ogni passaggio di fase della commessa."),
        ("Log aggiornamenti", "Traccia di chi modifica cosa: la disciplina del dato parte da qui."),
    ]
    ws.append(["Foglio", "Come usarlo"])
    _hdr(ws)
    for row in istruzioni:
        ws.append(list(row))
        ws.cell(ws.max_row, 2).alignment = Alignment(wrap_text=True)
    _widths(ws, (22, 90))
    ws.freeze_panes = "A2"
    mapping["istruzioni"] = {"sheet": ws.title, "range": f"A1:B{ws.max_row}"}

    # Fogli decisionali dal pacchetto consulenziale (#14): confronto opzioni + parametri
    # editabili. Il pack vive nel deliverable; no-op se assente.
    _pack = deliverable.get("consulenza_operativa")
    if isinstance(_pack, dict):
        _before = set(wb.sheetnames)
        append_decision_sheets(wb, _pack)
        for _name in wb.sheetnames:
            if _name not in _before:
                _ws = wb[_name]
                mapping[_name.lower().replace(" ", "_")] = {
                    "sheet": _name, "range": f"A1:H{_ws.max_row}"}

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Arial"
                cell.font = font

    wb.calculation.fullCalcOnLoad = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path, mapping


def excel_guide_section(mapping: dict) -> dict:
    """Sezione 'modello_excel' da iniettare nel deliverable → il PDF spiega i fogli
    (spec §11: quali fogli, come usarli, cosa si aggiorna da solo)."""
    fogli = [f"{m['sheet']}" for m in mapping.values()]
    return {
        "descrizione": ("Il report è accompagnato dal cruscotto Excel operativo: "
                        "è lo strumento di lavoro quotidiano, il PDF è la fotografia."),
        "fogli_inclusi": fogli,
        "come_usarlo": [
            "Registro commesse: una riga per commessa, stato solo dal Dizionario stati.",
            "KPI: i valori del report; le celle 'Da rilevare' vanno sostituite col dato reale.",
            "Dashboard: si aggiorna automaticamente dal foglio KPI, non scrivere a mano.",
            "RACI e Checklist: da compilare e validare con la direzione.",
            "Log aggiornamenti: tracciare ogni modifica.",
        ],
        "nota": "Nessuna cella contiene dati inventati: dove il dato manca trovi 'Da rilevare'.",
    }
