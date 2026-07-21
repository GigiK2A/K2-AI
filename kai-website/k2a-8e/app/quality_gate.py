"""Quality gate pre-consegna del report (spec §12).

Ultimo controllo prima di consegnare PDF+Excel: se rileva difetti BLOCCANTI il
report non va consegnato. Difetti oggettivi e domain-agnostici (nessun hardcoding
per singolo cliente/boost):

BLOCCANTI
- oggetti/involucri JSON non sballati (`{type,$value}`, `{value}`, JSON-stringato);
- `[object Object]`, `undefined`, repr di dict finiti come testo;
- dato mancante (N/D) con semaforo VERDE;
- KPI placeholder (valore=target=1) con semaforo verde.

NON BLOCCANTI (warning)
- stesso valore ripetuto con etichette diverse / KPI duplicati tra sezioni.

Restituisce un report tecnico leggibile: codice, severità, posizione, causa
probabile, correzione consigliata.
"""

from __future__ import annotations

from typing import Any, Iterator

from . import normalize as NORM

SEVERITY_BLOCK = "block"
SEVERITY_WARN = "warn"

# Valori che rappresentano "dato mancante" (mai un semaforo verde sopra).
_MISSING_TOKENS = {"", "n/d", "nd", "n.d.", "n/a", "na", "-", "—", "–",
                   "dati non disponibili", "non disponibile", "da rilevare", "da raccogliere"}


def _finding(code: str, severity: str, location: str, cause: str, fix: str,
             value: Any = None) -> dict:
    f = {"code": code, "severity": severity, "location": location,
         "cause": cause, "fix": fix}
    if value is not None:
        f["value"] = value
    return f


def _is_missing(v: Any) -> bool:
    v = NORM.unwrap_value(v)
    if v is None:
        return True
    if isinstance(v, str) and v.strip().lower() in _MISSING_TOKENS:
        return True
    return False


def _is_wrapper_dict(v: Any) -> bool:
    """True se v è un dict-involucro ({type,$value}/{value}) non ancora sballato."""
    return isinstance(v, dict) and not isinstance(NORM.unwrap_value(v), dict)


def _num(v: Any):
    v = NORM.unwrap_value(v)
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    return None


def _walk(obj: Any, path: str = "") -> Iterator[tuple[str, Any]]:
    """Percorre ricorsivamente la struttura, restituendo (path, valore)."""
    yield path, obj
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from _walk(v, f"{path}.{k}" if path else str(k))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from _walk(v, f"{path}[{i}]")


def _iter_kpis(deliverable: Any) -> Iterator[tuple[str, dict]]:
    """Trova i dict che 'sono' KPI: hanno un semaforo e un valore."""
    for path, node in _walk(deliverable):
        if isinstance(node, dict) and "semaforo" in node and (
                "valore" in node or "value" in node):
            yield path, node


def _kpi_value(kpi: dict) -> Any:
    return kpi.get("valore", kpi.get("value"))


def extract_kpis(deliverable: Any) -> list[dict]:
    """Estrae TUTTI i KPI del deliverable in forma normalizzata. È la fonte unica
    usata sia dal PDF che dal workbook Excel → coerenza per costruzione (spec §11).

    Ritorna: [{path, sezione, nome, valore, target, unita, semaforo, formula,
               fonte, nota}] — i campi assenti sono None (mai 0)."""
    out: list[dict] = []
    for path, kpi in _iter_kpis(deliverable):
        nome = NORM.to_text(kpi.get("nome") or kpi.get("label") or "").strip()
        if not nome:
            continue
        out.append({
            "path": path,
            "sezione": path.split(".")[0].split("[")[0],
            "nome": nome,
            "valore": NORM.unwrap_value(_kpi_value(kpi)),
            "target": NORM.unwrap_value(kpi.get("target")),
            "unita": NORM.to_text(kpi.get("unita") or kpi.get("unit") or "").strip() or None,
            "semaforo": NORM.to_text(kpi.get("semaforo")).strip().lower() or None,
            "formula": NORM.to_text(kpi.get("formula")).strip() or None,
            "fonte": NORM.to_text(kpi.get("fonte_dati") or kpi.get("fonte")
                                  or kpi.get("source") or "").strip() or None,
            "nota": NORM.to_text(kpi.get("nota")).strip() or None,
        })
    return out


def check_pdf_excel_coherence(deliverable: Any, xlsx_path: Any) -> list[dict]:
    """Test 7 (spec §11): gli stessi KPI devono avere valori/unità/fonti coerenti
    nel PDF e nell'Excel. Legge il foglio 'KPI' del workbook e confronta col
    deliverable. Ritorna findings (block su valore diverso, warn su KPI assente)."""
    findings: list[dict] = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path), data_only=False)
    except Exception as exc:
        return [_finding("xlsx_illeggibile", SEVERITY_BLOCK, str(xlsx_path),
                         f"workbook non apribile: {exc}",
                         "Rigenerare il workbook.")]
    if "KPI" not in wb.sheetnames:
        return [_finding("xlsx_senza_kpi", SEVERITY_BLOCK, str(xlsx_path),
                         "foglio 'KPI' assente dal workbook",
                         "Il workbook deve contenere il foglio KPI mappato dal report.")]
    ws = wb["KPI"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    by_name = {str(r[0]).strip().lower(): r for r in rows if r and r[0]}
    for kpi in extract_kpis(deliverable):
        row = by_name.get(kpi["nome"].lower())
        if row is None:
            findings.append(_finding(
                "kpi_assente_da_excel", SEVERITY_WARN, kpi["path"],
                f"KPI '{kpi['nome']}' presente nel PDF ma non nel foglio KPI",
                "Rigenerare il workbook dalla stessa estrazione KPI del PDF."))
            continue
        xl_val = row[1] if len(row) > 1 else None
        pdf_val = kpi["valore"]
        if pdf_val is None:
            continue          # nel PDF è 'dato mancante' → in Excel 'Da rilevare' (ok)
        if isinstance(pdf_val, (int, float)) and isinstance(xl_val, (int, float)):
            if abs(float(pdf_val) - float(xl_val)) > max(abs(float(pdf_val)) * 0.005, 0.01):
                findings.append(_finding(
                    "kpi_valore_incoerente", SEVERITY_BLOCK, kpi["path"],
                    f"KPI '{kpi['nome']}': PDF={pdf_val} vs Excel={xl_val}",
                    "PDF ed Excel devono derivare dalla stessa estrazione KPI."))
        elif str(xl_val).strip() != NORM.to_text(pdf_val).strip():
            findings.append(_finding(
                "kpi_valore_incoerente", SEVERITY_BLOCK, kpi["path"],
                f"KPI '{kpi['nome']}': PDF={NORM.to_text(pdf_val)!r} vs Excel={xl_val!r}",
                "PDF ed Excel devono derivare dalla stessa estrazione KPI."))
    return findings


def run_report_quality_gate(deliverable: Any, workbook: Any = None,
                            evidence: Any = None) -> dict:
    """Esegue il gate. Ritorna {ok, blocking, warnings, findings, report}."""
    findings: list[dict] = []

    # 1) involucri JSON non sballati (stringhe con pattern-leak o dict-involucro residui)
    for path, val in _walk(deliverable):
        if isinstance(val, str):
            for why in NORM.find_leaked_wrappers(val):
                findings.append(_finding(
                    "leaked_wrapper", SEVERITY_BLOCK, path, why,
                    "Passa il valore da normalize.to_text() prima del render.", val[:80]))
        elif _is_wrapper_dict(val) and path:   # il root può essere un dict legittimo
            findings.append(_finding(
                "wrapper_dict", SEVERITY_BLOCK, path,
                "valore-involucro {type,$value}/{value} non sballato",
                "Sballa con normalize.unwrap_value() a monte del render.",
                str(val)[:80]))

    # 2) KPI: dato mancante o placeholder con semaforo verde
    seen_kpis: list[tuple[str, Any, str]] = []
    for path, kpi in _iter_kpis(deliverable):
        sem = str(NORM.unwrap_value(kpi.get("semaforo")) or "").strip().lower()
        val = _kpi_value(kpi)
        tgt = kpi.get("target")
        label = str(NORM.unwrap_value(kpi.get("nome") or kpi.get("label") or "")).strip()
        if sem == "verde" and _is_missing(val):
            findings.append(_finding(
                "missing_green", SEVERITY_BLOCK, path,
                f"KPI '{label or path}' senza valore ma con semaforo verde",
                "Dato mancante ⇒ nessun semaforo (status=not_available, displayValue='Dati non disponibili').",
                label))
        elif sem == "verde" and _num(val) == 1 and _num(tgt) == 1:
            findings.append(_finding(
                "placeholder_green", SEVERITY_BLOCK, path,
                f"KPI '{label or path}' con valore=target=1 e semaforo verde (segnatura placeholder)",
                "Non riempire i KPI mancanti con 1/verde: marcali not_available.",
                label))
        # raccolta per dedup (warning)
        if label:
            seen_kpis.append((label.lower(), NORM.unwrap_value(val), path))

    # 2b) provenienza (spec §2/§3, Test 2): metriche che DICHIARANO una source vengono
    # validate contro l'evidence store. Le metriche senza source non sono toccate qui
    # (retro-compatibile → nessun falso positivo sui KPI attuali).
    if evidence is not None:
        from . import provenance as PROV
        for path, node in _walk(deliverable):
            if isinstance(node, dict) and node.get("source"):
                findings.extend(PROV.validate_metric(node, evidence, location=path))

    # 3) dedup (warning): stesso (etichetta, valore) ripetuto in più punti
    by_label: dict[str, list[tuple[Any, str]]] = {}
    for lbl, val, path in seen_kpis:
        by_label.setdefault(lbl, []).append((val, path))
    for lbl, occ in by_label.items():
        if len(occ) > 1 and len({str(v) for v, _ in occ}) == 1:
            findings.append(_finding(
                "duplicate_kpi", SEVERITY_WARN, ", ".join(p for _, p in occ),
                f"KPI '{lbl}' ripetuto identico in {len(occ)} punti",
                "Descrivi il KPI una sola volta; altrove richiamalo in forma sintetica (referencedIn)."))

    blocking = [f for f in findings if f["severity"] == SEVERITY_BLOCK]
    warnings = [f for f in findings if f["severity"] == SEVERITY_WARN]
    return {
        "ok": not blocking,
        "blocking": blocking,
        "warnings": warnings,
        "findings": findings,
        "report": format_report(findings),
    }


def format_report(findings: list[dict]) -> str:
    """Report tecnico leggibile del gate."""
    if not findings:
        return "Quality gate: nessun problema rilevato."
    lines = ["Quality gate — problemi rilevati:"]
    for f in findings:
        tag = "BLOCCANTE" if f["severity"] == SEVERITY_BLOCK else "warning"
        lines.append(f"[{tag}] {f['code']} @ {f['location']}")
        lines.append(f"    causa: {f['cause']}")
        lines.append(f"    correzione: {f['fix']}")
    return "\n".join(lines)
