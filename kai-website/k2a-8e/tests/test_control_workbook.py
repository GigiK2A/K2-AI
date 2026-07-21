"""Test del cruscotto Excel ControlBoost (spec §11) e coerenza PDF/Excel (Test 7).

I KPI del workbook vengono dalla STESSA estrazione del PDF (quality_gate.extract_kpis)
→ qui si verifica la coerenza per costruzione, la gestione dei dati mancanti
('Da rilevare', mai 0) e che il check di coerenza becchi un workbook manomesso.
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import control  # noqa: E402
from app.quality_gate import check_pdf_excel_coherence, extract_kpis  # noqa: E402
from app.xlsx import DA_RILEVARE, excel_guide_section, render_control_workbook  # noqa: E402

# Stessi input operativi usati da tests/test_control.py → KPI deterministici reali.
_FORM = {
    "mese": "giugno", "anno": 2026, "azienda": "K2A S.r.l.",
    "fatturato": 100000, "costi_operativi": 73000,
    "crediti_clienti": 145000, "cash_flow_mese": 8000,
    "clienti_attivi": 42, "clienti_persi": 2, "clienti_nuovi": 5,
    "ore_fatturabili": 1210, "ore_lavorate": 1600,
    "progetti_in_corso": 12, "progetti_in_ritardo": 3,
}

_SHEETS_ATTESI = {
    "Registro commesse", "Dizionario stati", "Registro blocchi", "KPI", "Dashboard",
    "Piano 30-60-90", "RACI", "Checklist", "Log aggiornamenti", "Istruzioni",
}


def _build(tmp: Path):
    deliverable, meta = control.apply_controlboost({"meta": {}}, _FORM)
    assert meta is not None, "il binder deve calcolare i KPI dai dati operativi"
    path, mapping = render_control_workbook(deliverable, _FORM, tmp / "cruscotto.xlsx")
    return deliverable, path, mapping


def test_workbook_has_all_ten_sheets():
    with tempfile.TemporaryDirectory() as td:
        _, path, mapping = _build(Path(td))
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert set(wb.sheetnames) == _SHEETS_ATTESI
        # ogni voce del mapping punta a un foglio esistente
        for section, m in mapping.items():
            assert m["sheet"] in wb.sheetnames, section
            assert ":" in m["range"]


def test_kpi_sheet_matches_pdf_extraction():
    """Test 7 spec: stessi KPI, stessi valori nei due file."""
    with tempfile.TemporaryDirectory() as td:
        deliverable, path, _ = _build(Path(td))
        findings = check_pdf_excel_coherence(deliverable, path)
        assert findings == [], findings


def test_coherence_check_catches_tampered_value():
    with tempfile.TemporaryDirectory() as td:
        deliverable, path, _ = _build(Path(td))
        from openpyxl import load_workbook
        wb = load_workbook(path)
        wb["KPI"].cell(2, 2).value = 999999   # manomissione: valore diverso dal PDF
        wb.save(path)
        findings = check_pdf_excel_coherence(deliverable, path)
        assert any(f["code"] == "kpi_valore_incoerente" and f["severity"] == "block"
                   for f in findings)


def test_missing_kpi_sheet_blocks():
    with tempfile.TemporaryDirectory() as td:
        deliverable, path, _ = _build(Path(td))
        from openpyxl import load_workbook
        wb = load_workbook(path)
        del wb["KPI"]
        wb.save(path)
        findings = check_pdf_excel_coherence(deliverable, path)
        assert any(f["code"] == "xlsx_senza_kpi" for f in findings)


def test_missing_data_is_da_rilevare_never_zero():
    """Un KPI col valore mancante → 'Da rilevare' in cella, mai 0 (spec §11)."""
    deliverable = {"kpi_processi": [
        {"nome": "Commesse in ritardo", "valore": 22, "target": 10,
         "unita": "%", "semaforo": "rosso"},
        {"nome": "Tempo medio chiusura", "valore": None, "target": None, "semaforo": None},
    ]}
    with tempfile.TemporaryDirectory() as td:
        path, _ = render_control_workbook(deliverable, {}, Path(td) / "c.xlsx")
        from openpyxl import load_workbook
        rows = {r[0]: r for r in load_workbook(path)["KPI"].iter_rows(min_row=2, values_only=True)}
        assert rows["Commesse in ritardo"][1] == 22
        # NB: extract_kpis richiede il campo semaforo per riconoscere un KPI;
        # il secondo ha semaforo=None ma presente → estratto, valore 'Da rilevare'
        assert rows["Tempo medio chiusura"][1] == DA_RILEVARE
        assert 0 not in [rows["Tempo medio chiusura"][1], rows["Tempo medio chiusura"][3]]


def test_owner_and_formula_columns_present():
    """Ogni KPI porta formula/fonte/periodicità/owner/aggiornamento (spec §11)."""
    with tempfile.TemporaryDirectory() as td:
        _, path, _ = _build(Path(td))
        from openpyxl import load_workbook
        header = [c.value for c in load_workbook(path)["KPI"][1]]
        for col in ("Formula", "Fonte", "Periodicità", "Owner", "Ultimo aggiornamento"):
            assert col in header


def test_extract_kpis_shared_source():
    deliverable, _ = control.apply_controlboost({"meta": {}}, _FORM)
    kpis = extract_kpis(deliverable)
    names = {k["nome"] for k in kpis}
    assert "Fatturato" in names or len(names) >= 4   # le 4 prospettive BSC estratte
    assert all(k["valore"] is not None for k in kpis)  # il binder scarta i None


def test_guide_section_lists_sheets():
    with tempfile.TemporaryDirectory() as td:
        _, _, mapping = _build(Path(td))
        guide = excel_guide_section(mapping)
        assert "KPI" in guide["fogli_inclusi"]
        assert guide["come_usarlo"]
        assert "inventat" in guide["nota"]
