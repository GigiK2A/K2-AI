from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app import grounding, quality  # noqa: E402


def json_dumps(obj) -> str:
    return json.dumps(obj, ensure_ascii=False)
from app.xlsx import render_finance_workbook  # noqa: E402


def _k2a_input():
    return {
        "ragione_sociale": "K2A S.r.l.s.",
        "settore_ateco": "71.12",
        "n_dipendenti": 4,
        "bilanci": [{
            "anno": 2024,
            "ricavi": 789766.50,
            "utile_netto": 87688.18,
            "totale_attivo": 648622.39,
            "capitale_sociale": 200.00,
            "riserve": 0.0,
            "utili_portati_nuovo": 88239.26,
            "risultato_esercizio": 87688.18,
            "passivita_verso_terzi": 472494.95,
            "debiti_bancari": 51951.38,
            "mutui_finanziamenti": 40907.33,
            "imposte": 8522.00,
            "oneri_finanziari": 13181.28,
            "ammortamenti_svalutazioni": 12470.37,
        }],
    }


def test_finance_reconciliation_regression_k2a():
    out, errors, notes = quality.normalize_finance_inputs(_k2a_input())
    assert not errors
    b = out["bilanci"][0]
    assert b["patrimonio_netto"] == 176127.44
    assert b["ebitda"] == 121861.83
    assert b["debiti_finanziari"] == 92858.71
    assert any("PN=" in n for n in notes)


def test_finance_wrong_equity_is_corrected_when_two_quadratures_agree():
    inp = _k2a_input()
    inp["bilanci"][0]["patrimonio_netto"] = 87688.18
    out, errors, notes = quality.normalize_finance_inputs(inp)
    assert not errors
    assert out["bilanci"][0]["patrimonio_netto"] == 176127.44
    assert any("corretto" in n for n in notes)


def test_finance_ambiguous_equity_still_blocks():
    inp = _k2a_input()
    b = inp["bilanci"][0]
    b["patrimonio_netto"] = 87688.18
    b.pop("passivita_verso_terzi")
    _, errors, _ = quality.normalize_finance_inputs(inp)
    assert any("patrimonio netto incoerente" in e for e in errors)


def test_four_section_printed_liabilities_are_not_treated_as_debt():
    inp = _k2a_input()
    b = inp["bilanci"][0]
    b.pop("passivita_verso_terzi")
    b["totale_passivita_prima_risultato"] = 560934.21
    out, errors, _ = quality.normalize_finance_inputs(inp)
    assert not errors
    assert out["bilanci"][0]["passivita_verso_terzi"] == 472494.95
    assert out["bilanci"][0]["patrimonio_netto"] == 176127.44


def test_missing_identity_blocks_all_reports():
    _, errors, _ = quality.prepare_inputs(
        "flusso-strategyboost-pmi",
        {"type": "object", "required": ["obiettivo"], "properties": {"obiettivo": {"type": "string"}}},
        {"obiettivo": "crescita"},
    )
    assert any("ragione_sociale" in e for e in errors)


def test_placeholder_and_unsupported_money_block():
    bad = {
        "meta": {"azienda": "Acme SRL"},
        "analisi": "L'azienda opera a [città]. Risparmio previsto €30000 e +25% EBITDA.",
    }
    findings = grounding.integrity_findings(
        bad, inputs={"ragione_sociale": "Acme SRL"}, facts={}, citazioni=[])
    codes = {f["code"] for f in findings if f["severity"] == "block"}
    assert "placeholder_leak" in codes     # [città] è template trapelato → block (strict=True default)
    assert "numero_non_grounded" in codes  # €30000 / EBITDA non in input → block


def test_advisory_mode_allows_sector_benchmarks_not_hard_financials():
    """strict=False (boost qualitativo):
    - CTR/traffico % benchmark → warn, NON block (scenario corretto: WebBoost)
    - €/EBITDA hard-financial → block OVUNQUE (sicurezza)
    - [città] placeholder specifico → block OVUNQUE (output rotto)
    """
    d = {
        "meta": {"azienda": "Acme SRL"},
        "seo": "Tasso di CTR medio +5% e conversioni +8% nel settore.",
        "proiezione": "Risparmio stimato €40000 EBITDA anno.",
        "area": "Studio a [città] operante.",
    }
    inp = {"ragione_sociale": "Acme SRL", "url": "https://acme.it"}
    findings = grounding.integrity_findings(d, inputs=inp, facts={}, citazioni=[], strict=False)
    blocks = {f["code"] for f in findings if f["severity"] == "block"}
    warns  = {f["code"] for f in findings if f["severity"] == "warn"}
    # benchmark soft: numero_non_grounded ma NON hard-financial → warn
    assert "numero_non_grounded" in warns or len(blocks) == 0 or "numero_non_grounded" not in blocks
    # €/EBITDA hard: resta block anche su qualitativo
    assert "numero_non_grounded" in blocks
    # [città] placeholder specifico: block sempre
    assert "placeholder_leak" in blocks


def test_assumption_label_allows_scenario_number():
    good = {
        "meta": {"azienda": "Acme SRL"},
        "scenario": "Scenario illustrativo, assunzione da validare: ricavi +10%.",
    }
    findings = quality.unsupported_number_findings(good, {"ragione_sociale": "Acme SRL"}, {})
    assert not findings


def test_qualitative_allows_illustrative_number_with_honest_markers():
    """Scelta utente: i boost qualitativi (StrategyBoost) POSSONO includere numeri
    illustrativi SE marcati come ipotesi esplicita. I marker onesti più naturali
    ('(ipotesi)', 'a titolo illustrativo', 'da confermare') devono esentare, così il
    modello non viene bloccato quando etichetta onestamente una proiezione."""
    exempt = {
        "meta": {"azienda": "Acme SRL"},
        "piano": "Budget marketing indicativo 10.000€ a titolo illustrativo (ipotesi da confermare).",
        "roi": "ROI atteso del 30% (ipotesi esplicita, da confermare con dati reali).",
    }
    assert not quality.unsupported_number_findings(exempt, {"ragione_sociale": "Acme SRL"}, {}, strict=False)


def test_naked_hard_financial_still_blocks_on_qualitative():
    """Protezione intatta: un numero hard-financial NUDO (senza marker di ipotesi)
    resta bloccato anche sui qualitativi — non si spacciano cifre inventate per fatti."""
    naked = {"meta": {"azienda": "Acme SRL"}, "piano": "Investimento di 50.000€ con ROI del 30%."}
    blocks = [f for f in quality.unsupported_number_findings(naked, {"ragione_sociale": "Acme SRL"}, {}, strict=False)
              if f["severity"] == "block"]
    assert blocks


def test_scrub_template_placeholders_neutralises_leaks():
    """Il deep-gen a volte lascia [città]/[regione] quando non conosce la sede (nonostante
    il prompt lo vieti) → il gate li BLOCCA (placeholder_leak, refuse reale visto in prod:
    job_f1d31380ff0d). Lo scrub deterministico li neutralizza PRIMA del gate: il report si
    consegna invece di fallire. [nome]→cliente reale, luoghi ignoti→neutro."""
    d = {
        "meta": {"azienda": "Studio Evolution"},
        "intro": "Lo studio [nome] con sede a [città], in [regione], opera dal [mese/anno].",
        "voci": ["Espansione in [città]", "Presidio [regione]"],
    }
    out = quality.scrub_template_placeholders(d, {"ragione_sociale": "Studio Evolution"})
    blob = json_dumps(out)
    # nessun segnaposto a bracket sopravvive
    assert "[città]" not in blob and "[regione]" not in blob and "[nome]" not in blob and "[mese/anno]" not in blob
    # [nome] riempito col cliente reale
    assert "Studio Evolution" in blob
    # e il gate ora NON blocca più (placeholder_leak sparito)
    leaks = [f for f in grounding.integrity_findings(out, inputs={"ragione_sociale": "Studio Evolution"},
                                                     facts={}, citazioni=[], strict=False)
             if f["code"] == "placeholder_leak" and f["severity"] == "block"]
    assert not leaks


def test_scrub_preserves_uppercase_markers_and_refs():
    """Lo scrub NON deve toccare i marker legittimi maiuscoli ([IPOTESI]) né i ref numerici."""
    d = {"x": "Stima [IPOTESI] da confermare; vedi nota [1]."}
    out = quality.scrub_template_placeholders(d, {"ragione_sociale": "Acme"})
    blob = json_dumps(out)
    assert "[IPOTESI]" in blob and "[1]" in blob


def test_financial_strict_not_loosened_by_weak_estimate_word():
    """FinanceBoost (strict): 'stimato €40000 EBITDA' resta block — i numeri finanziari
    vanno grounded/bound, non 'stimati'. La parola debole 'stima' NON esenta."""
    d = {"meta": {"azienda": "Acme SRL"}, "x": "Risparmio stimato €40000 EBITDA anno."}
    blocks = [f for f in quality.unsupported_number_findings(d, {"ragione_sociale": "Acme SRL"}, {}, strict=True)
              if f["severity"] == "block"]
    assert blocks


def _voci_input():
    return {
        "ragione_sociale": "Esempio S.r.l.",
        "settore_ateco": "70.22",
        "n_dipendenti": 5,
        "bilanci": [{
            "anno": 2024,
            "voci": [
                {"sezione": "attivo", "descrizione": "BANCHE C/C", "importo": 100000.0},
                {"sezione": "attivo", "descrizione": "CLIENTI", "importo": 50000.0},
                {"sezione": "attivo", "descrizione": "IMPIANTI E MACCHINARI", "importo": 30000.0},
                {"sezione": "passivo", "descrizione": "CAPITALE SOCIALE", "importo": 10000.0},
                {"sezione": "passivo", "descrizione": "RISULTATI PORTATI A NUOVO", "importo": 20000.0},
                {"sezione": "passivo", "descrizione": "FONDI AMMORTAMENTO IMPIANTI E MACCHINARI", "importo": 5000.0},
                {"sezione": "passivo", "descrizione": "MUTUI E FINANZIAMENTI", "importo": 40000.0},
                {"sezione": "passivo", "descrizione": "FORNITORI", "importo": 80000.0},
                {"sezione": "ricavi", "descrizione": "RICAVI DELLE VENDITE", "importo": 200000.0},
                {"sezione": "costi", "descrizione": "COSTI DELLA PRODUZIONE", "importo": 175000.0},
                {"sezione": "risultato", "descrizione": "UTILE DEL PERIODO", "importo": 25000.0},
            ],
        }],
    }


def test_voci_transcription_feeds_reconciliation():
    # Le voci grezze trascritte vengono riclassificate deterministicamente e popolano
    # gli aggregati che la riconciliazione poi valida: niente più PN/EBITDA stimati dall'LLM.
    out, errors, notes = quality.normalize_finance_inputs(_voci_input())
    assert not errors
    b = out["bilanci"][0]
    assert b["patrimonio_netto"] == 55000.0
    assert b["passivita_verso_terzi"] == 120000.0  # fondi ammortamento NON contati come debito
    assert b["debiti_finanziari"] == 40000.0
    assert b["attivo_corrente"] == 150000.0  # derivato (prima restava vuoto → Excel vuoto)
    assert any("riclassificato" in n for n in notes)


def test_voci_that_do_not_balance_are_blocked():
    inp = _voci_input()
    inp["bilanci"][0]["voci"].append({"sezione": "attivo", "descrizione": "CASSA", "importo": 99999.0})
    _, errors, _ = quality.normalize_finance_inputs(inp)
    assert any("quadr" in e.lower() for e in errors)


def test_finance_xlsx_has_live_formulas(tmp_path):
    normalized, errors, _ = quality.normalize_finance_inputs(_k2a_input())
    assert not errors
    path = render_finance_workbook(normalized, tmp_path / "finance.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == ["Input bilancio", "Indici", "Scenari"]
    assert wb["Indici"]["B2"].value.startswith("=IFERROR(")
    assert wb["Scenari"]["D2"].value.startswith("=IF(")
    assert wb["Scenari"]["B2"].value is None
