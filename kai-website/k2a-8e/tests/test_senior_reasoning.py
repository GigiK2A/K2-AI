"""Test spec §15 — da generatore di report a consulente senior.

Caso di riferimento: FinanceBoost liquidità (18 dipendenti, 1,9M fatturato,
145k incassi/mese, 158k uscite/mese, 95k scoperto, 18k interessi, 72% top3,
DSO 75). Verifica: ogni dato usato, catene causa-effetto, confronto soluzioni,
calcoli automatici, simulazioni, raccomandazioni col perché/come, soglie sempre
classificate, PDF/Excel coerenti.
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import consulting, decision, insight, reasoning, scenario  # noqa: E402

_CASE = {
    "azienda": "Impresa Test S.r.l.", "dipendenti": 18,
    "fatturato_annuo": 1_900_000,
    "incassi_mese": 145_000, "uscite_mese": 158_000,
    "scoperto": 95_000, "interessi_annui": 18_000,
    "concentrazione_top3": 72, "concentrazione_top1": 40,
    "dso": 75,
    "contesto": "Tensione di liquidità: fatturazione in ritardo, incassi a 70-80 "
                "giorni, scoperto bancario stabile e interessi in crescita. "
                "Serve capire come sistemare la cassa e i pagamenti.",
}


def _pack():
    return consulting.build_finance_pack(_CASE, {})


# ── §15.1: ogni dato importante usato almeno una volta ─────────────────────────
def test_all_important_data_used():
    insights, facts = insight.derive_finance_insights(_CASE)
    cov = insight.coverage_report(facts)
    # dopo l'intero pack anche dso/dpo/concentrazione entrano nelle analisi
    pack = _pack()
    assert pack["copertura_dati"]["dati_non_sfruttati"] == [], \
        pack["copertura_dati"]["dati_non_sfruttati"]
    assert pack["copertura_dati"]["copertura_pct"] == 100.0


# ── §15.4 (§5 spec): calcoli automatici dai dati ───────────────────────────────
def test_quantitative_insights_computed():
    insights, _ = insight.derive_finance_insights(_CASE)
    by_id = {i["id"]: i for i in insights}
    # 145k − 158k = −13k €/mese
    assert by_id["cash.saldo_mensile"]["valore"] == -13_000
    # 18k / 95k ≈ 18.9% costo effettivo
    assert abs(by_id["debt.costo_scoperto"]["valore"] - 18.9) < 0.2
    # capitale nei crediti: 1.9M/365×75 ≈ 390k
    assert abs(by_id["wc.capitale_in_crediti"]["valore"] - 390_411) < 1000
    # concentrazione
    assert by_id["risk.concentrazione"]["valore"] == 72
    # ogni insight ha formula + dati usati + spiegazione (il PERCHÉ)
    for i in insights:
        assert i["formula"] and i["dati_usati"] and len(i["spiegazione"]) > 40
        assert i["source"] == "system_calculated" and i["confidence"] == "A"


# ── §15.2: almeno una catena causa-effetto, ben formata ────────────────────────
def test_causal_chains_built_and_valid():
    insights, _ = insight.derive_finance_insights(_CASE)
    chains = reasoning.build_finance_chains(insights, _CASE)
    assert len(chains) >= 2                     # spirale scoperto + concentrazione
    for c in chains:
        assert reasoning.validate_chain(c) == [], reasoning.validate_chain(c)
        fasi = [n["fase"] for n in c["catena"]]
        assert "osservazione" in fasi and "cause" in fasi
        assert "conseguenze" in fasi and "intervento" in fasi
        # l'osservazione è ancorata agli insight (evidenze)
        oss = next(n for n in c["catena"] if n["fase"] == "osservazione")
        assert oss["evidenze"]


def test_no_chains_without_data():
    insights, _ = insight.derive_finance_insights({"azienda": "X"})
    assert insights == []
    assert reasoning.build_finance_chains(insights, {}) == []


# ── §15.5 (§6 spec): forecast 13 settimane × 3 scenari ─────────────────────────
def test_forecast_three_scenarios_13_weeks():
    fc = scenario.cash_forecast_13w(_CASE)
    assert fc is not None
    assert set(fc["scenari"]) == {"prudente", "realistico", "critico"}
    for nome, sc in fc["scenari"].items():
        assert len(sc["settimane"]) == 13
        for r in sc["settimane"]:
            assert set(r) == {"settimana", "entrate", "uscite", "saldo"}
    # coerenza aritmetica: realistico parte da −95k e perde ~3k/settimana
    reale = fc["scenari"]["realistico"]["settimane"]
    assert reale[0]["saldo"] < -95_000
    assert reale[12]["saldo"] < reale[0]["saldo"]
    # ipotesi esplicite (spec: "se mancano dati, esplicitare le ipotesi")
    assert len(fc["ipotesi"]) >= 4


def test_no_forecast_without_monthly_data():
    assert scenario.cash_forecast_13w({"fatturato_annuo": 1_000_000}) is None


# ── §15.6 (§9 spec): simulazioni what-if ───────────────────────────────────────
def test_what_if_simulations():
    sims = scenario.what_if(_CASE)
    assert len(sims) >= 3
    domande = " ".join(s["domanda"] for s in sims).lower()
    assert "30 giorni" in domande       # top client in ritardo
    assert "dso" in domande             # DSO +15
    assert "10%" in domande             # calo fatturato
    for s in sims:
        assert s["calcolo"] and s["dati_usati"] and s["risultato"]


# ── §15.3 (§7 spec): confronto soluzioni con conclusione motivata ──────────────
def test_solution_comparison_three_options():
    insights, _ = insight.derive_finance_insights(_CASE)
    conf = decision.finance_options(_CASE, insights)
    assert len(conf["opzioni"]) >= 3
    for o in conf["opzioni"]:
        for k in ("vantaggi", "svantaggi", "costi", "rischi", "tempi",
                  "complessita", "dipendenze", "quando_sceglierla", "quando_evitarla"):
            assert o.get(k), f"{o['opzione']}: manca {k}"
    concl = conf["conclusione_motivata"].lower()
    assert "perché" in concl or "perche" in concl   # spiega la preferenza


# ── §15.7 (§8/§11 spec): raccomandazioni con perché e come ─────────────────────
def test_recommendations_answer_why_and_how():
    insights, _ = insight.derive_finance_insights(_CASE)
    recs = decision.finance_recommendations(_CASE, insights)
    assert recs
    for r in recs:
        for k in ("perche", "perche_ora", "perche_questa", "perche_non_altre"):
            assert r.get(k), f"{r['id']}: manca {k}"
        op = r["operativo"]
        for k in ("chi", "quando", "con_quali_dati", "cadenza", "validazione",
                  "kpi_generati", "decisore"):
            assert op.get(k), f"{r['id']}: manca operativo.{k}"


# ── §15.8 (§12 spec): nessuna soglia senza classificazione ─────────────────────
def test_thresholds_always_classified():
    insights, _ = insight.derive_finance_insights(_CASE)
    recs = decision.finance_recommendations(_CASE, insights)
    for r in recs:
        for s in r["soglie"]:
            assert s["classificazione"] in ("dato_aziendale", "benchmark",
                                            "best_practice", "proposta_iniziale",
                                            "ipotesi")
    import pytest
    with pytest.raises(AssertionError):
        decision.soglia("45 giorni", "numero_magico")


# ── planner: il caso liquidità viene riconosciuto e il pack passa il gate ───────
def test_classifier_and_pack_pass_quality_gate():
    assert consulting.classify_problem(_CASE) == "finanza_liquidita"
    from app import provenance as PROV
    from app.quality_gate import run_report_quality_gate
    pack = _pack()
    deliverable = {"executive_summary": "Caso liquidità.", "consulenza_operativa": pack}
    res = run_report_quality_gate(deliverable, evidence=PROV.build_evidence(_CASE))
    assert res["ok"] is True, res["report"]


# ── §15.9 (§10 spec): Excel vivo e coerente col PDF ────────────────────────────
def test_excel_treasury_sheets_and_coherence():
    from openpyxl import load_workbook
    from app.xlsx import render_finance_workbook
    pack = _pack()
    with tempfile.TemporaryDirectory() as td:
        path = render_finance_workbook(_CASE, Path(td) / "m.xlsx", pack=pack)
        wb = load_workbook(path)
        for sheet in ("Forecast 13 settimane", "Aging crediti", "Piano incassi",
                      "KPI tesoreria", "Simulazioni", "Piano azione",
                      "Registro decisioni"):
            assert sheet in wb.sheetnames, sheet
        # KPI tesoreria = stessi valori degli insight del PDF (coerenza)
        kpi_rows = {r[0]: r[1] for r in
                    wb["KPI tesoreria"].iter_rows(min_row=2, values_only=True)}
        by_titolo = {i["titolo"]: i["valore"] for i in pack["insight_derivati"]}
        for titolo, valore in by_titolo.items():
            assert kpi_rows.get(titolo) == valore, titolo
        # forecast: cella input incassi = dato fornito, formule vive presenti
        fw = wb["Forecast 13 settimane"]
        assert fw["B2"].value == 145_000
        assert str(fw["D7"].value).startswith("=")


def test_excel_without_bilanci_needs_pack():
    import pytest
    from app.xlsx import render_finance_workbook
    with tempfile.TemporaryDirectory() as td:
        with pytest.raises(ValueError):
            render_finance_workbook({"x": 1}, Path(td) / "no.xlsx")   # compat invariata


# ── PDF end-to-end col pacchetto finance ───────────────────────────────────────
def test_pdf_renders_finance_pack():
    from app.render import render_generic_pdf
    pack = _pack()
    deliverable = {"executive_summary": "Tensione di liquidità strutturale.",
                   "consulenza_operativa": pack}
    with tempfile.TemporaryDirectory() as td:
        pdf = Path(td) / "out.pdf"
        render_generic_pdf(deliverable, {"nome": "FinanceBoost"}, [], pdf)
        content = pdf.read_bytes()
        assert content[:5] == b"%PDF-"
        assert len(content) > 25_000
        try:
            import pdfplumber
            with pdfplumber.open(str(pdf)) as doc:
                text = "\n".join(p.extract_text() or "" for p in doc.pages)
            assert "OSSERVAZIONE" in text and "CAUSE" in text      # catena causale
            assert "13 settimane" in text                          # forecast
            assert "Quando evitarla" in text                       # confronto opzioni
            assert "Perché ora" in text                            # 4 perché
        except ImportError:
            pass
