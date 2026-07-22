"""Motore di composizione premium (review deliverable): struttura per-caso, pagina-diagnosi
CEO, matrice decisionale, perché-non, raccomandazione finale, KPI governance, niente N/D,
Excel operativo. Tutto grounded: i componenti esistono SOLO se il pack ha i dati.
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import compose, consulting  # noqa: E402
from app import styling as ST  # noqa: E402

_S = ST.styles()

# caso diagnostico (efficienza): ipotesi pesate + escluse + KPI
_EFF = {"contesto": ("Utili diminuiti ma fatturato stabile. Aumento di riunioni, livelli di "
                     "approvazione, revisioni interne. Dipendenti meno produttivi. Cambio del "
                     "responsabile operativo pochi mesi prima. Nessuna nuova assunzione, nessun "
                     "aumento salariale. Materie prime stabili."),
        "delta_fatturato_pct": 2, "costo_personale_delta_pct": 12, "spese_generali_delta_pct": 18}

# caso decisionale (marketing canali): opzioni + conclusione
_MKT = {"contesto": "dipendenza dal canale principale, riequilibrare canali di vendita e marketing",
        "canale_principale_pct": 70, "fatturato": 900000}


def _deliv(case, skill):
    return {"executive_summary": "x",
            "consulenza_operativa": consulting.build_pack(skill, case, {})}


def _flat(flowables) -> str:
    out = []
    for f in flowables:
        for attr in ("text",):
            if hasattr(f, attr):
                out.append(str(getattr(f, attr)))
        if hasattr(f, "_cellvalues"):
            stack = [f]
            while stack:
                t = stack.pop()
                for row in t._cellvalues:
                    for cell in row:
                        items = cell if isinstance(cell, (list, tuple)) else [cell]
                        for it in items:
                            if hasattr(it, "text"):
                                out.append(str(it.text))
                            elif hasattr(it, "_cellvalues"):
                                stack.append(it)
    return " ".join(out)


# ── richiesta finale: la struttura VARIA per caso (niente template fisso) ────────────────
def test_structure_differs_by_case():
    p1 = compose.build_report_plan(_deliv(_EFF, "cruscotto-direzionale"))
    p2 = compose.build_report_plan(_deliv(_MKT, "campaign-plan"))
    assert p1 != p2
    assert p1["diagnosis_one_pager"] and p1["kpi_governance"]      # diagnosi → one-pager+KPI
    assert p2["decision_matrix"] and not p2["kpi_governance"]      # decisionale → matrice


def test_empty_deliverable_renders_nothing_premium():
    plan = compose.build_report_plan({"executive_summary": "solo testo"})
    assert not any(v for k, v in plan.items() if not k.startswith("_"))
    assert compose.premium_front({"executive_summary": "x"}, _S) == []
    assert compose.premium_back({"executive_summary": "x"}, _S) == []


# ── #2 pagina-diagnosi CEO ───────────────────────────────────────────────────────────────
def test_one_pager_has_ceo_fields():
    txt = _flat(compose.diagnosis_one_pager(_deliv(_EFF, "cruscotto-direzionale"), _S)).upper()
    for frag in ("PROBLEMA PRINCIPALE", "CAUSA PIÙ PROBABILE", "DECISIONE RACCOMANDATA",
                 "URGENZA", "RISCHIO COMPLESSIVO"):
        assert frag in txt, frag


# ── #3 fatti/evidenze/ipotesi con confidenza ─────────────────────────────────────────────
def test_evidence_ledger_confidence():
    txt = _flat(compose.evidence_ledger(_deliv(_EFF, "cruscotto-direzionale"), _S))
    assert "IPOTESI" in txt
    assert "Alta" in txt or "Media" in txt or "Bassa" in txt


# ── #6/#7 filosofia diagnostica: ipotesi da verificare + incertezza dichiarata ───────────
def test_evidence_ledger_has_to_verify_and_preliminary():
    txt = _flat(compose.evidence_ledger(_deliv(_EFF, "cruscotto-direzionale"), _S))
    assert "ANCORA DA VERIFICARE" in txt                 # #6: ipotesi da verificare
    assert "PRELIMINARE" in txt.upper()                  # #7: incertezza dichiarata
    assert "confermate, ridimensionate o escluse" in txt


def test_generation_prompt_diagnostic_honesty():
    # il prompt di generazione 8e porta le regole della review
    from app.llm import _QUALITA_TRASVERSALE as Q
    assert "NIENTE ENTITÀ INVENTATE" in Q                # #4 no ruoli/tool inventati
    assert "PMO" in Q and "Power BI" in Q
    assert "CITA LE PAROLE DEL CLIENTE" in Q             # #3 citare il cliente
    assert "PROPORZIONATE ALLA DIMENSIONE" in Q          # #5 azioni proporzionate
    assert "non è un fatto" in Q.lower()                 # #2 deduzioni ≠ fatti
    assert "INCERTEZZA DICHIARATA" in Q                  # #7


# ── #9 matrice decisionale + #8 perché-non + #7 raccomandazione finale ───────────────────
def test_decision_matrix_and_why_not():
    d = _deliv(_MKT, "campaign-plan")
    m = _flat(compose.decision_matrix(d, _S))
    assert "Consigliata" in m or "Alternativa" in m
    w = _flat(compose.why_not_section(d, _S))
    assert "Perché non" in w
    f = _flat(compose.final_recommendation(d, _S))
    assert "decisione consigliata" in f.lower()
    assert "cambiare raccomandazione" in f.lower() or "Cosa ci farebbe" in f


# ── #10 KPI governance ───────────────────────────────────────────────────────────────────
def test_kpi_governance_operational_columns():
    txt = _flat(compose.kpi_governance(_deliv(_EFF, "cruscotto-direzionale"), _S))
    for frag in ("KPI", "Frequenza", "Responsabile", "fuori soglia"):
        assert frag in txt, frag


# ── #6 niente N/D nel PDF ────────────────────────────────────────────────────────────────
def test_nd_never_rendered():
    from app.render import _rich, _scalar_str
    assert _scalar_str("N/D") == "Parametro da definire"
    assert _scalar_str(None) == "Parametro da definire"
    assert "N/D" not in _rich("Il margine è N/D per ora.")
    assert "da definire" in _rich("Il margine è N/D per ora.")


# ── #1 exec summary: perché prima del cosa (nei diagnostici) ─────────────────────────────
def test_exec_summary_reasoning_order():
    from app.render import _exec_summary
    txt = _flat(_exec_summary(_deliv(_EFF, "cruscotto-direzionale"), _S))
    assert "La causa più probabile" in txt


# ── PDF end-to-end col Livello 3 ─────────────────────────────────────────────────────────
def test_pdf_contains_decision_level():
    from app.render import render_generic_pdf
    d = _deliv(_EFF, "cruscotto-direzionale")
    with tempfile.TemporaryDirectory() as td:
        pdf = Path(td) / "o.pdf"
        render_generic_pdf(d, {"nome": "ControlBoost"}, [], pdf)
        assert pdf.read_bytes()[:5] == b"%PDF-"
        try:
            import pdfplumber
            with pdfplumber.open(str(pdf)) as doc:
                text = "\n".join(p.extract_text() or "" for p in doc.pages).upper()
            assert "DIAGNOSI IN 60 SECONDI" in text
            assert "RACCOMANDAZIONE FINALE" in text
            assert "N/D" not in text
        except ImportError:
            pass


# ── #14 Excel operativo: fogli decisionali ───────────────────────────────────────────────
def test_xlsx_decision_sheets():
    from openpyxl import Workbook, load_workbook
    from app.xlsx import append_decision_sheets
    pack = consulting.build_pack("campaign-plan", _MKT, {})
    wb = Workbook(); wb.remove(wb.active)
    append_decision_sheets(wb, pack)
    assert "Confronto opzioni" in wb.sheetnames
    ws = wb["Confronto opzioni"]
    assert ws["H1"].value == "Decisione (compila)"
    assert ws.max_row >= 3           # header + >=2 opzioni
    # no-op senza pack
    wb2 = Workbook(); wb2.remove(wb2.active)
    append_decision_sheets(wb2, None)
    assert not wb2.sheetnames
