"""Excel 'modello vivo' dal deliverable 8e (2° file del bundle): ogni lista-di-oggetti
del deliverable diventa un foglio editabile. Deterministico, niente rete/crediti."""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.lib.xlsx_renderer import render_deliverable_8e_xlsx  # noqa: E402

_DELIV = {
    "meta": {"cliente": "Studio Associato Evolution"},
    "opzioni_strategiche": {"opzioni": [
        {"opzione": "Posizionamento ibrido", "attrattivita": 5, "fattibilita": 3, "rischio": 3},
        {"opzione": "Focalizzazione rinnovabili", "attrattivita": 4, "fattibilita": 4, "rischio": 4},
    ]},
    "piano_strategico": {"iniziative": [
        {"titolo": "Ridisegno sito dual-track", "priorita": "Alta", "tempi": "3 mesi"},
        {"titolo": "LinkedIn B2B", "priorita": "Media", "tempi": "continuo"},
    ]},
    "executive_summary": "testo discorsivo che NON deve diventare una tabella",
}


def test_deliverable_diventa_excel_multifoglio():
    from openpyxl import load_workbook
    data = render_deliverable_8e_xlsx(_DELIV)
    assert len(data) > 1000
    wb = load_workbook(io.BytesIO(data))
    # Sintesi + un foglio per ogni lista-di-oggetti (opzioni, iniziative)
    assert "Sintesi" in wb.sheetnames
    assert len(wb.sheetnames) >= 3
    names = " ".join(wb.sheetnames).lower()
    assert "opzioni" in names and "iniziative" in names


def test_nessun_crash_su_valori_annidati():
    # un valore dict/list dentro una cella non deve far crashare (json fallback)
    deliv = {"meta": {"cliente": "X"}, "sez": {"righe": [{"a": {"nested": 1}, "b": [1, 2]}]}}
    data = render_deliverable_8e_xlsx(deliv)
    assert len(data) > 800


if __name__ == "__main__":
    import pytest
    sys.exit(pytest.main([__file__, "-q"]))
