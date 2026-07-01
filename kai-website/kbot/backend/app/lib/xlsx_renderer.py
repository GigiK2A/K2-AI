"""Render a report-design-system analysis JSON to an .xlsx workbook.

Reuses the SAME `{meta, blocks[]}` payload produced by
`lib.analysis.generate_analysis_json` and rendered to PDF by `lib.pdf_renderer`.
Each `data_table` block becomes a real spreadsheet sheet (ideal for editorial
calendars, content pillars, financial projections); textual blocks
(executive_summary, narrative, conclusions, two_column) are summarised on a
leading "Sintesi" sheet.

Kept deliberately defensive: the LLM JSON shape can vary, so every field access
is guarded and unknown block types degrade to text instead of raising.
"""
from __future__ import annotations

import io
import json
import re
from typing import Any, Dict, List

K2_GREEN = "0C7A6F"
K2_DARK = "063B36"
_MAX_SHEET_NAME = 31  # Excel hard limit


def _strip_html(value: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value if value is not None else ""))
    text = text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    return re.sub(r"\s+", " ", text).strip()


_EURO_RE = re.compile(r"^€\s*(-?[\d.]+(?:,\d+)?)$")
_PCT_RE = re.compile(r"^(-?[\d.]+(?:,\d+)?)\s*%$")
_NUM_RE = re.compile(r"^-?[\d.]+(?:,\d+)?$")


def _typed_cell(value: Any) -> tuple[Any, str | None]:
    """Mantiene numeri/formule come tali; niente workbook composto solo da testo."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value, '#,##0.00;[Red](#,##0.00);-'
    text = _strip_html(value)
    if text.startswith("="):
        return text, None
    for rx, fmt, scale in ((_EURO_RE, '€ #,##0.00;[Red](€ #,##0.00);-', 1),
                           (_PCT_RE, '0.0%', 100), (_NUM_RE, '#,##0.00;[Red](#,##0.00);-', 1)):
        m = rx.match(text)
        if not m:
            continue
        raw = m.group(1) if m.lastindex else text
        try:
            n = float(raw.replace(".", "").replace(",", ".")) / scale
            return n, fmt
        except ValueError:
            pass
    return text, None


def _safe_sheet_title(title: str, used: set) -> str:
    name = re.sub(r"[\[\]:*?/\\]", " ", _strip_html(title) or "Foglio").strip() or "Foglio"
    name = name[:_MAX_SHEET_NAME]
    base, i = name, 2
    while name.lower() in used:
        suffix = f" {i}"
        name = (base[: _MAX_SHEET_NAME - len(suffix)] + suffix).strip()
        i += 1
    used.add(name.lower())
    return name


def _summary_lines(blocks: List[dict]) -> List[str]:
    """Flatten textual blocks into readable lines for the Sintesi sheet."""
    lines: List[str] = []
    for b in blocks:
        if not isinstance(b, dict):
            continue
        btype = b.get("type") or ""
        title = _strip_html(b.get("title") or b.get("heading") or "")
        if btype == "executive_summary":
            if title:
                lines.append(title.upper())
            score = b.get("score")
            if score not in (None, ""):
                lines.append(f"Punteggio: {score}")
            body = _strip_html(b.get("body") or b.get("summary"))
            if body:
                lines.append(body)
            lines.append("")
        elif btype in ("narrative", "callout", "text"):
            if title:
                lines.append(title.upper())
            body = _strip_html(b.get("body") or b.get("body_html"))
            if body:
                lines.append(body)
            lines.append("")
        elif btype == "two_column":
            for side in ("left", "right"):
                col = b.get(side) or {}
                if isinstance(col, dict):
                    h = _strip_html(col.get("heading"))
                    if h:
                        lines.append(h.upper())
                    items = col.get("items") or []
                    if isinstance(items, list):
                        for it in items:
                            lines.append(f"  • {_strip_html(it)}")
                    body = _strip_html(col.get("body") or col.get("body_html"))
                    if body:
                        lines.append(body)
            lines.append("")
        elif btype == "section_break":
            lbl = _strip_html(b.get("title") or b.get("layer") or "")
            if lbl:
                lines.append("")
                lines.append(f"=== {lbl.upper()} ===")
        elif btype == "executive_dashboard":
            lines.append((title or "Cruscotto direzionale").upper())
            gauge = b.get("gauge") or {}
            if gauge.get("value") not in (None, ""):
                lines.append(f"Punteggio generale: {gauge.get('value')}/{gauge.get('max') or 100}")
            status = b.get("status") or {}
            if status.get("label"):
                lines.append(f"Stato: {_strip_html(status.get('label'))}")
            for sub in b.get("subscores") or []:
                if isinstance(sub, dict) and sub.get("label"):
                    lines.append(f"  • {_strip_html(sub.get('label'))}: {sub.get('value')}")
            for p in b.get("problems") or []:
                lines.append(f"  ▸ Criticità: {_strip_html(p)}")
            for o in b.get("opportunities") or []:
                lines.append(f"  ▸ Opportunità: {_strip_html(o)}")
            verdict = b.get("verdict") or {}
            if verdict.get("text"):
                dec = _strip_html(verdict.get("decision") or "")
                lines.append(f"Verdetto: {_strip_html(verdict.get('text'))}" + (f" → {dec}" if dec else ""))
            lines.append("")
        elif btype == "financial_impact":
            lines.append((title or "Impatto economico").upper())
            for key, dflt in (("inaction", "Se non agisci"), ("action", "Se agisci")):
                d = b.get(key) or {}
                if isinstance(d, dict) and (d.get("value") or d.get("note")):
                    lbl = _strip_html(d.get("label") or dflt)
                    val = _strip_html(d.get("value") or "")
                    note = _strip_html(d.get("note") or "")
                    lines.append(f"  • {lbl}: {val}" + (f" — {note}" if note else ""))
            lines.append("")
        elif btype == "recommendations":
            lines.append((title or "Azioni raccomandate").upper())
            for h in b.get("horizons") or []:
                if not isinstance(h, dict):
                    continue
                lines.append(_strip_html(h.get("label") or ""))
                for it in h.get("items") or []:
                    lines.append(f"  • {_strip_html(it)}")
            lines.append("")
        elif btype == "decision_board":
            lines.append((title or "Decision Board").upper())
            for c in b.get("cells") or b.get("items") or []:
                if isinstance(c, dict):
                    lines.append(f"  • {_strip_html(c.get('label'))}: {_strip_html(c.get('value'))}")
            lines.append("")
        elif btype == "source_legend":
            lines.append((title or "Affidabilità dei dati").upper())
            for it in b.get("items") or []:
                if isinstance(it, dict):
                    tag = _strip_html(it.get("tag") or "")
                    txt = _strip_html(it.get("text") or "")
                    note = _strip_html(it.get("note") or "")
                    lines.append(f"  [{tag}] {txt}" + (f" — {note}" if note else ""))
            lines.append("")
        elif btype == "conclusions":
            lines.append(_strip_html(b.get("title") or "Conclusioni e prossimi passi").upper())
            left = b.get("left") or {}
            if isinstance(left, dict):
                if left.get("heading"):
                    lines.append(_strip_html(left.get("heading")))
                body = _strip_html(left.get("body_html") or left.get("body"))
                if body:
                    lines.append(body)
            right = b.get("right") or {}
            if isinstance(right, dict):
                if right.get("heading"):
                    lines.append(_strip_html(right.get("heading")))
                for m in right.get("milestones") or []:
                    if isinstance(m, dict):
                        label = _strip_html(m.get("label"))
                        items = ", ".join(_strip_html(i) for i in (m.get("items") or []))
                        lines.append(f"  • {label}: {items}" if items else f"  • {label}")
            lines.append("")
    return [ln for ln in lines]


def _as_table_block(b: dict) -> dict | None:
    """Normalizza benchmark_table / severity_matrix in {columns, rows} per un foglio."""
    btype = b.get("type")
    if btype == "benchmark_table":
        cols = b.get("columns") or ["KPI", "Azienda", "Settore", "Delta"]
        rows = []
        for r in b.get("rows") or []:
            if isinstance(r, dict):
                rows.append([r.get("kpi") or r.get("label") or "", r.get("company") or r.get("azienda") or "",
                             r.get("sector") or r.get("settore") or "", r.get("delta") or ""])
            elif isinstance(r, list):
                rows.append(r)
        return {"title": b.get("title") or "Benchmark", "columns": cols[:4], "rows": rows}
    if btype == "severity_matrix":
        rows = []
        for it in b.get("items") or b.get("rows") or []:
            if isinstance(it, dict):
                rows.append([it.get("problem") or it.get("title") or it.get("label") or "",
                             it.get("severity") or it.get("level") or "", it.get("effort") or "", it.get("roi") or ""])
        return {"title": b.get("title") or "Matrice priorità",
                "columns": ["Problema", "Severity", "Effort", "ROI"], "rows": rows}
    return None


def render_xlsx(analysis: Dict[str, Any]) -> bytes:
    """Return .xlsx bytes for a `{meta, blocks[]}` analysis payload."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    meta = analysis.get("meta") or {}
    blocks = [b for b in (analysis.get("blocks") or []) if isinstance(b, dict)]
    title = _strip_html(meta.get("title") or "Report K2-AI") or "Report K2-AI"

    header_fill = PatternFill("solid", fgColor=K2_GREEN)
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color=K2_DARK)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    used_names: set = set()

    # --- Sintesi sheet (always first) ---
    ws = wb.active
    ws.title = _safe_sheet_title("Sintesi", used_names)
    ws["A1"] = title
    ws["A1"].font = title_font
    row = 3
    for line in _summary_lines(blocks):
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = wrap
        row += 1
    ws.column_dimensions["A"].width = 110

    # --- data_table / benchmark_table / severity_matrix blocks → one sheet each ---
    table_count = 0
    for raw_b in blocks:
        if raw_b.get("type") == "data_table":
            b = raw_b
        else:
            b = _as_table_block(raw_b)
            if b is None:
                continue
        table_count += 1
        columns = [_strip_html(c) for c in (b.get("columns") or [])]
        rows = b.get("rows") or []
        sheet_name = _safe_sheet_title(b.get("title") or f"Tabella {table_count}", used_names)
        tws = wb.create_sheet(sheet_name)

        tws.cell(row=1, column=1, value=_strip_html(b.get("title") or sheet_name)).font = title_font
        header_row = 3
        ncols = max(len(columns), max((len(r) for r in rows if isinstance(r, list)), default=1))

        for ci in range(1, ncols + 1):
            label = columns[ci - 1] if ci - 1 < len(columns) else ""
            cell = tws.cell(row=header_row, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        for ri, r in enumerate(rows, start=header_row + 1):
            cells = r if isinstance(r, list) else [r]
            for ci in range(1, ncols + 1):
                value = cells[ci - 1] if ci - 1 < len(cells) else ""
                typed, number_format = _typed_cell(value)
                cell = tws.cell(row=ri, column=ci, value=typed)
                if number_format:
                    cell.number_format = number_format
                cell.alignment = wrap
                cell.border = border

        # Column widths from longest cell (capped), then freeze the header.
        for ci in range(1, ncols + 1):
            longest = len(columns[ci - 1]) if ci - 1 < len(columns) else 10
            for r in rows:
                if isinstance(r, list) and ci - 1 < len(r):
                    longest = max(longest, len(_strip_html(r[ci - 1])))
            tws.column_dimensions[get_column_letter(ci)].width = min(max(longest + 2, 12), 50)
        tws.freeze_panes = tws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===================== Bundle Boost: Excel "modello vivo" =====================
# I blueprint dell'8e chiedono per ogni Boost un bundle multi-file (docx report +
# xlsx "modello vivo" + html dashboard). Qui rendiamo il 2° file: un Excel editabile
# dal deliverable 8e — ogni lista-di-oggetti (opzioni scorate, iniziative, forze
# Porter, KPI) diventa un foglio. Riusa render_xlsx (già testato).
def _cell_text(v: Any) -> str:
    if isinstance(v, (dict, list)):
        return _strip_html(json.dumps(v, ensure_ascii=False))
    return _strip_html(v)


def _tables_from_deliverable(obj: Any, path: str, out: List[dict]) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            _tables_from_deliverable(v, (path + " " + str(k)).strip(), out)
    elif isinstance(obj, list) and obj and all(isinstance(x, dict) for x in obj):
        cols: List[str] = []
        for x in obj:
            for k in x.keys():
                if k not in cols:
                    cols.append(k)
        rows = [[_cell_text(x.get(c, "")) for c in cols] for x in obj]
        out.append({"type": "data_table", "title": (path or "Dati").replace("_", " "),
                    "columns": [c.replace("_", " ") for c in cols], "rows": rows})


def render_deliverable_8e_xlsx(deliverable: Dict[str, Any], *, titolo: str = "") -> bytes:
    """Excel 'modello vivo' da un deliverable 8e (il 2° artefatto del bundle).
    Ogni lista-di-oggetti del deliverable → un foglio editabile. Deterministico."""
    blocks: List[dict] = []
    _tables_from_deliverable(deliverable, "", blocks)
    meta = {"title": titolo or (deliverable.get("meta") or {}).get("cliente") or "Modello operativo K2-AI"}
    return render_xlsx({"meta": meta, "blocks": blocks})
